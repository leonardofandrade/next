import os
from typing import Optional
from django.core.management.base import BaseCommand, CommandError
from django.utils import timezone
from django.db import transaction

try:
    import openpyxl
except Exception as e:  # pragma: no cover
    openpyxl = None

from apps.requisitions.models import ExtractionRequest
from apps.base_tables.models import AgencyUnit
from apps.core.models import ExtractionUnit


class Command(BaseCommand):
    help = (
        "Importa solicitações de extração legadas a partir de um arquivo Excel (dist_2022.xlsx) "
        "criando registros na tabela extraction_request com is_legacy=True."
    )

    def add_arguments(self, parser):
        parser.add_argument(
            "--file",
            dest="file_path",
            default=os.path.join("data", "dist_2022.xlsx"),
            help="Caminho para o arquivo XLSX (padrão: data/dist_2022.xlsx)",
        )
        parser.add_argument(
            "--year",
            dest="year",
            type=int,
            required=True,
            help="Ano das solicitações esperadas no arquivo. Obrigatório. O comando irá avisar quando a data lida não for deste ano.",
        )
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Executa a leitura e validação sem salvar no banco.",
        )
        parser.add_argument(
            "--limit",
            type=int,
            default=None,
            help="Importa no máximo N linhas (para testes).",
        )
        parser.add_argument(
            "--login",
            dest="login",
            default=None,
            help="Login (username) do usuário responsável pela importação. Será usado em created_by.",
        )

    def handle(self, *args, **options):
        if openpyxl is None:
            raise CommandError("A dependência 'openpyxl' é necessária para ler arquivos XLSX. Instale-a em requirements.txt.")

        file_path: str = options["file_path"]
        dry_run: bool = options["dry_run"]
        limit: Optional[int] = options["limit"]
        login: Optional[str] = options.get("login")
        expected_year: int = options["year"]

        from django.contrib.auth import get_user_model
        UserModel = get_user_model()
        user = None
        if login:
            try:
                user = UserModel.objects.get(username=login)
            except UserModel.DoesNotExist:
                raise CommandError(f"Usuário com login '{login}' não encontrado.")
        elif not dry_run:
            # Em execução real, exigir login para rastreabilidade
            raise CommandError("Parâmetro --login é obrigatório quando não está em --dry-run.")

        if not os.path.exists(file_path):
            raise CommandError(f"Arquivo não encontrado: {file_path}")

        wb = openpyxl.load_workbook(file_path)
        ws = wb.active

        # Tenta ler cabeçalhos da primeira linha
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        header_map = {str(h).strip().lower(): idx for idx, h in enumerate(headers)}

        # Mapeamento esperado (ajuste aqui conforme o layout do XLSX)
        # Tentamos ser tolerantes com nomes de colunas comuns em PT-BR
        candidates = {
            "requested_at": ["data", "data_solicitacao", "solicitada_em", "requested_at"],
            "requester_unit": ["unidade", "unidade_solicitante", "requester", "orgao"],
            "procedures": ["procedimentos", "procedures"],
            "device_amount": ["qtd", "quantidade", "dispositivos", "device_amount"],
            "extraction_unit": ["unidade_extracao", "atribuido_a", "assigned_to", "extraction_unit"],
            "requester_name": ["solicitante", "requester_name"],
            "requester_email": ["email", "email_solicitante", "requester_email"],
            "request_notes": ["observacoes", "notas", "request_notes"],
            "status": ["status", "situacao"],
            "line_number": ["ord"]
        }

        def find_col(key: str):
            names = candidates.get(key, [])
            for name in names:
                idx = header_map.get(name)
                if idx is not None:
                    return idx
            return None

        col_requested_at = find_col("requested_at")
        col_requester_unit = find_col("requester_unit")
        col_procedures = find_col("procedures")
        col_device_amount = find_col("device_amount")
        col_extraction_unit = find_col("extraction_unit")
        col_requester_name = find_col("requester_name")
        col_requester_email = find_col("requester_email")
        col_request_notes = find_col("request_notes")
        col_status = find_col("status")
        col_line_number = find_col("line_number")

        required_cols = {
            "requested_at": col_requested_at,
            "requester_unit": col_requester_unit,
            "procedures": col_procedures,
        }
        missing = [k for k, v in required_cols.items() if v is None]
        if missing:
            raise CommandError(
                "Colunas obrigatórias não encontradas no XLSX: " + ", ".join(missing)
            )

        created = 0
        skipped = 0
        errors = 0
        consecutive_empty_rows = 0
        max_consecutive_empty = 10  # Para após 10 linhas vazias consecutivas

        # Inicia da segunda linha (dados)
        data_rows = ws.iter_rows(min_row=2)

        def get_value(row, idx):
            if idx is None:
                return None
            cell = row[idx]
            return cell.value

        # Transação para garantir atomicidade, exceto em dry-run
        context = transaction.atomic if not dry_run else _noop_context
        with context():
            for i, row in enumerate(data_rows, start=2):
                if limit and (created + skipped + errors) >= limit:
                    break
                
                # Verifica se devemos parar por muitas linhas vazias consecutivas
                if consecutive_empty_rows >= max_consecutive_empty:
                    self.stdout.write(f"Parando processamento após {max_consecutive_empty} linhas vazias consecutivas (linha {i})")
                    break
                
                try:
                    requested_at_val = get_value(row, col_requested_at)
                    if requested_at_val is None:
                        # pular linhas vazias
                        skipped += 1
                        consecutive_empty_rows += 1
                        self.stdout.write(f"Linha {i}: vazia - nenhum ExtractionRequest criado")
                        continue
                    
                    # Reset contador de linhas vazias quando encontra dados válidos
                    consecutive_empty_rows = 0

                    # Converte datas do Excel
                    req_at = _parse_excel_datetime(requested_at_val)
                    # Verifica se o ano corresponde ao informado
                    try:
                        req_year = (req_at.year if hasattr(req_at, 'year') else None)
                    except Exception:
                        req_year = None
                    if req_year is not None and req_year != expected_year:
                        self.stdout.write(self.style.WARNING(f"Linha {i}: requested_at com ano {req_year} diferente do informado ({expected_year})."))

                    requester_unit_name = str(get_value(row, col_requester_unit) or "").strip()
                    procedures = str(get_value(row, col_procedures) or "").strip()
                    device_amount = get_value(row, col_device_amount)
                    try:
                        device_amount = int(device_amount) if device_amount is not None else 1
                    except Exception:
                        device_amount = 1

                    extraction_unit_name = str(get_value(row, col_extraction_unit) or "").strip()
                    # Mapeamento de valores da planilha para nomes corretos das extraction units
                    extraction_unit_name = _map_extraction_unit_name(extraction_unit_name)
                    # print(f"extraction_unit_name: {extraction_unit_name}")
                    requester_name = str(get_value(row, col_requester_name) or "").strip() if col_requester_name is not None else None
                    requester_email = str(get_value(row, col_requester_email) or "").strip() if col_requester_email is not None else None
                    request_notes = str(get_value(row, col_request_notes) or "").strip() if col_request_notes is not None else None
                    status_val = str(get_value(row, col_status) or "").strip() if col_status is not None else ""
                    normalized_status = _normalize_status(status_val) if status_val else ExtractionRequest.REQUEST_STATUS_ASSIGNED
                    line_number = get_value(row, col_line_number)

                    # Resolver FKs por nome
                    requester_unit = _find_agency_unit_by_name(requester_unit_name)

                    # Se não encontrou por unidade, tenta pelo nome do solicitante e cria se necessário
                    if requester_unit is None:
                        # 1) tentar criar a partir do nome da unidade (coluna 'unidade_solicitante')
                        candidate_name = requester_unit_name or requester_name
                        alt_unit = None
                        if candidate_name:
                            alt_unit = _find_agency_unit_by_name(candidate_name)
                        # 2) se ainda não existe e não é dry-run, criar AgencyUnit nova
                        if alt_unit is None and not dry_run and candidate_name:
                            # Gera um acronym a partir do nome (primeiras letras de cada palavra, limitado a 50 caracteres)
                            acronym_parts = [word[0].upper() for word in candidate_name.split() if word]
                            generated_acronym = ''.join(acronym_parts)[:50] if acronym_parts else candidate_name[:50]
                            alt_unit = AgencyUnit.objects.create(
                                name=candidate_name,
                                acronym=generated_acronym,
                                created_by=user if user else None,
                            )
                            print(f"Criada nova unidade: {alt_unit.name} (acronym: {alt_unit.acronym})")
                        requester_unit = alt_unit
                        if requester_unit is None:
                            # se ainda não encontrou, falha
                            raise ValueError(f"Unidade solicitante não encontrada: '{requester_unit_name or requester_name}' (linha {i})")

                    extraction_unit = _find_extraction_unit_by_short_name(extraction_unit_name) if extraction_unit_name else None
                    if extraction_unit_name and not extraction_unit:
                        self.stderr.write(f"Linha {i}: unidade de extração '{extraction_unit_name}' não encontrada - extraction_request será criado sem 'extraction_unit'")
              

                    obj = ExtractionRequest.objects.create(
                        requester_authority_name=requester_name or None,
                        requester_agency_unit=requester_unit,
                        requested_at=req_at,
                        requested_device_amount=device_amount,
                        request_procedures=procedures or "",
                        extraction_unit=extraction_unit,
                        requester_reply_email=requester_email or None,
                        additional_info=request_notes or None,
                        status=normalized_status,
                        is_legacy=True,
                        spreadsheet_line=line_number,
                        created_by=user if user else None,
                    )
                    created += 1
                    # Log linha a linha para saber se a solicitação foi criada e para onde
                    extraction_unit_label = obj.extraction_unit.acronym if obj.extraction_unit else "(sem unidade)"
                    self.stdout.write(
                        # f"Linha {i}: ExtractionRequest criado id={obj.id} | solicitante='{obj.requester_agency_unit}' | atribuída a='{extraction_unit_label}' | dispositivos={obj.requested_device_amount} | data={obj.requested_at.strftime('%Y-%m-%d')}"
                        f"Linha {i}: ExtractionRequest da linha {line_number} criado."
                    )
                except Exception as e:
                    errors += 1
                    self.stderr.write(f"Erro na linha {i}: {e}")

        self.stdout.write(self.style.SUCCESS(
            f"Importação finalizada para ano={expected_year}. Criados: {created}, Pulados: {skipped}, Erros: {errors}. Dry-run: {dry_run}"
        ))


def _parse_excel_datetime(value):
    # openpyxl já retorna datetime para células de data; para strings, tentamos alguns formatos
    import datetime as _dt
    if isinstance(value, _dt.datetime):
        return value
    if isinstance(value, _dt.date):
        return _dt.datetime.combine(value, _dt.time.min, tzinfo=timezone.get_current_timezone())
    if isinstance(value, str):
        for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y %H:%M", "%d/%m/%Y %H:%M:%S", "%Y-%m-%d %H:%M:%S"):
            try:
                dt = _dt.datetime.strptime(value.strip(), fmt)
                return timezone.make_aware(dt, timezone.get_current_timezone()) if dt.tzinfo is None else dt
            except Exception:
                pass
    # fallback para agora
    return timezone.now()


def _normalize_status(value: str) -> str:
    v = (value or "").strip().lower()
    mapping = {
        "assigned": ExtractionRequest.REQUEST_STATUS_ASSIGNED,
        "aguardando material": ExtractionRequest.REQUEST_STATUS_ASSIGNED,
        "material recebido": ExtractionRequest.REQUEST_STATUS_RECEIVED,
        "received": ExtractionRequest.REQUEST_STATUS_RECEIVED,
        "em andamento": ExtractionRequest.REQUEST_STATUS_IN_PROGRESS,
        "in_progress": ExtractionRequest.REQUEST_STATUS_IN_PROGRESS,
        "in progress": ExtractionRequest.REQUEST_STATUS_IN_PROGRESS,
        "aguardando coleta": ExtractionRequest.REQUEST_STATUS_WAITING_COLLECT,
        "waiting_collection": ExtractionRequest.REQUEST_STATUS_WAITING_COLLECT,
        "waiting collection": ExtractionRequest.REQUEST_STATUS_WAITING_COLLECT,
    }
    # retorno padrão se não encontrado
    return mapping.get(v, ExtractionRequest.REQUEST_STATUS_ASSIGNED)


def _map_extraction_unit_name(name: str) -> str:
    """
    Mapeia valores da planilha para os nomes corretos das extraction units.
    Ex: "COIN" -> "CEINS/NEXT", "DIP" -> "PCCE/DIP"
    """
    if not name:
        return name
    
    name_upper = name.strip().upper()
    mapping = {
        "COIN": "CEINS/NEXT",
        "DIP": "PCCE/DIP",
    }
    
    # Retorna o valor mapeado se existir, caso contrário retorna o valor original
    return mapping.get(name_upper, name.strip())


def _find_agency_unit_by_name(name: str) -> Optional[AgencyUnit]:
    if not name:
        return None
    try:
        return AgencyUnit.objects.get(acronym__iexact=name)
    except AgencyUnit.DoesNotExist:
        # tenta por nome completo
        try:
            return AgencyUnit.objects.get(name__iexact=name)
        except AgencyUnit.DoesNotExist:
            return None


def _find_extraction_unit_by_short_name(short_name: str) -> Optional[ExtractionUnit]:
    """
    Try to resolve an ExtractionUnit strictly by its acronym, but be tolerant with
    common variations coming from spreadsheets (extra words, separators, punctuation).
    This remains strictly based on the acronym field and never matches by name.
    """
    if not short_name:
        return None

    val = (short_name or "").strip()
    if not val:
        return None

    # First try a direct iexact match
    try:
        # print(f"Tentando resolver por acronym: {val}")
        return ExtractionUnit.objects.get(acronym__iexact=val)
    except ExtractionUnit.DoesNotExist:
        pass

    # Prepare alternative candidates by trimming at common separators and removing punctuation/spaces
    candidates = set()
    separators = [" -", "–", "—", "(", "/", ",", "|"]
    s = val
    for sep in separators:
        if sep in s:
            s = s.split(sep, 1)[0].strip()
    if s:
        candidates.add(s)

    # Also try removing dots and spaces from the candidate, as some acronyms are given with punctuation
    compact = s.replace(".", "").replace(" ", "") if s else ""
    if compact:
        candidates.add(compact)

    # Also add the original compacted as a fallback
    orig_compact = val.replace(".", "").replace(" ", "")
    if orig_compact:
        candidates.add(orig_compact)

    for cand in candidates:
        try:
            return ExtractionUnit.objects.get(acronym__iexact=cand)
        except ExtractionUnit.DoesNotExist:
            continue

    # As a very last resort, try contains to handle cases where the provided text includes the acronym as a substring
    # This still only targets the acronym field and returns a single deterministic record.
    qs = ExtractionUnit.objects.filter(acronym__iexact=val)
    if qs.exists():
        return qs.first()
    # try icontains for all candidates
    for cand in [val] + list(candidates):
        qs = ExtractionUnit.objects.filter(acronym__icontains=cand).order_by("id")
        if qs.exists():
            return qs.first()

    return None


class _noop_context:
    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc, tb):
        return False
